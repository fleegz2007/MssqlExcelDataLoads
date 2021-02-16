import openpyxl
import config
from openpyxl import load_workbook

def extractExcel(filepath):
    datalistich = []
    wb = load_workbook(filepath, data_only=True)
    ws = wb.active
    minrow = ws.min_row
    maxrow = ws.max_row
    mincol = ws.min_column
    maxcol = ws.max_column
    print("Beginning interchange extraction from Excel document...")
    for i in range (mincol, maxcol+1):
        datalistich.append([])
        for j in range (minrow+1, maxrow):
            #Unmute and use this if you want to set conditions of dataload if ws.cell(row=j, column=i).value == 'condition1' or ws.cell(row=j, column=i).value == 'condition2' or ws.cell(row=j, column=i).value == 'condition3':
                datalistich[i-1].append(ws.cell(row=j, column=i).value)
    return datalistich


def colNames(filepath):
    wb = load_workbook(filepath, data_only=True)
    ws = wb.active
    mincol = ws.min_column
    maxcol = ws.max_column
    colnames = []
    for i in range (mincol, maxcol+1):
        colnames.append(ws.cell(row=1, column=i).value)
    return colnames